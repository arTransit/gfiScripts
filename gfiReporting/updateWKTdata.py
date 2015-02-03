"""
updateWKTdata.py

Complete west kootenay ridership & revenue spreadsheet for given month.

"""


import sys
import os
import argparse
import datetime
import openpyxl
import cx_Oracle

# list of routes to include in query
ROUTELIST = (
        781,782,783,784,785,786,787,786,787,788,789,790,
        791,792,795,909,910,906,907,914,911,912,913,915
        )

# target tab in xlsx file - not this will need to be changed each year
TARGETWORKSHEET='14-15'

# Location id for Trail
LOC_N=11


"""
getArgs()

use argparse to get command line args
    year, month: data of data
    file: name of WestKootenay report spreadsheet (file will be overwritten)
    credentials: user id, password, database for oracle connection

"""
def getArgs():
    argsPsr = argparse.ArgumentParser(description='Update Ridership & Revenue Report')
    argsPsr.add_argument('-y','--year',required=True,type=int,help='eg 2014')
    argsPsr.add_argument('-m','--month',required=True,type=int,help='eg 12')
    argsPsr.add_argument('-f','--file',required=True,help='eg file.xlsx')
    argsPsr.add_argument('-c','--credentials',required=True,help='user/pass@GFI')
    args = argsPsr.parse_args()
    args.error = False
    if (args.year > datetime.date.today().year) or (args.year < 2000):
        print "ERROR: year out of range (2000 - %d)" % datetime.date.today().year
        args.error = True
    if (args.month > 12) or (args.month < 1):
        print "ERROR: month out of range (1 - 12)"
        args.error = True
    return args


"""
gfiQuery()
Generate array of ridership and revenue for each route for the given
month and year.
    Arguments:
        credentials: userid, password, database
        year, month: date of data
    Returns:
        associative array of ridership & revenue for each route (key)
        each value is an associate array with 'revenue', 'ridership' keys
        string values return by db query converted to float

"""
def gfiQuery(credentials,year,month,location):
    sql = (
        "select route,curr_r,rdr_c "
        "from mrtesum "
        "where "
            "loc_n=%s and "
            "mday=to_date('%s%s01','YYYYMMDD') and "
            "route in ( %s ) "
        "order by route "
        ) % (
                str(location), 
                str(args.year), 
                ('000' + str(args.month) )[-2:], 
                ','.join([str(x) for x in ROUTELIST]) 
                )
    #print "sql: " + sql

    try:
        connection = cx_Oracle.connect(credentials)
    except cx_Oracle.DatabaseError:
        connection.close()
        status = False
        return

    try:
        cursor = connection.cursor()
        cursor.execute(sql)
    except cx_Oracle.DatabaseError:
        connection.close()
        status = False
        return

    data={}
    headers = [i[0].lower() for i in cursor.description]
    for r in cursor: 
        #print r
        data[r[0]] = {'revenue':float("{:.2f}".format(float(r[1]))), 'ridership':int(r[2])}
    connection.close()

    return data


"""
intWithoutError()
    convert given value to int. If it is a string or other value return 0.

"""
def intWithoutError(x):
    try: return int(x)
    except: return 0


"""
testFile()
    Check if spreadsheet is valid.

    Tests if requested month/column is empty and if spreadsheet 'looks' like the standard format

"""
def testFile(year,month,filename):
    wb = openpyxl.load_workbook(filename=filename)
    if TARGETWORKSHEET not in wb.get_sheet_names():
        print "ERROR: cannot find target worksheet " + TARGETWORKSHEET
        return False

    ws = wb.get_sheet_by_name( TARGETWORKSHEET )
    targetColumn= ((month -4) %12) +3  # convert month to column
    if sum([intWithoutError(x.value) for x in ws.columns[targetColumn][:30] ]) > 0: 
        print "ERROR: data exists in month " + str(month)
        #print sum([intWithoutError(x.value) for x in ws.columns[targetColumn] ])
        #print ws.columns[targetColumn]
        return False

    if ws['A5'].value != 'Route Code':
        print "ERROR: cell A5 should say _Route Code_"
        return False

    return True


"""
updateSpreadsheet()
    Write results of database query to new spreadsheet.

    Route numbers in column A provide index to data array.

"""
def updateSpreadsheet(data,filename,year,month):
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.get_sheet_by_name( TARGETWORKSHEET )
    targetColumn= ((month -4) %12) +4  # convert month to spreadsheet column
    #print "appendData: targetColumn:" + str(targetColumn)

    for r in ws.rows:
        routeCode = intWithoutError( r[0].value )
        #print "appendData: routeCode:" + str(routeCode)
        rRow = int( r[0].row )
        #print "appendData: rRow:" + str(rRow)
        if routeCode in data.keys():
            if rRow <39: 
                ws.cell(row=rRow,column=targetColumn).value = data[routeCode]['ridership']
                #print "appendData: data ridership:" + str(data[routeCode]['ridership'])
            else: 
                ws.cell(row=rRow,column=targetColumn).value = data[routeCode]['revenue']
                #print "appendData: data revenue:" + str(data[routeCode]['revenue'])

    wb.save(os.path.splitext(filename)[0]+'_new'+os.path.splitext(filename)[1])


"""
createReport()
    Main route to call query and generate report, handles errors

"""
def createReport(location,year,month,filename,credentials):

    gq = gfiQuery(credentials,year,month,location)
    if not len(gq):
        print "ERROR: no data received"
        print gq
        sys.exit(1)

    if testFile(year,month, filename): updateSpreadsheet(gq,filename,year,month)


if __name__ == '__main__':
    args = getArgs()
    if args.error:
        print "Not completed."
        sys.exit(1)

    createReport(LOC_N,args.year,args.month,args.file,args.credentials)

    print "Completed."
    sys.exit(0)




