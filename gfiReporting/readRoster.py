#
# readRoster.py
#
# Read the given Conventional Bus Roster and generate
# an sqlite db used to process the data
#

import openpyxl
import sqlite3
import sys
import os
import argparse



class rostertable():
    """
    Class representing xlsx verion of bus roster
     - the input.
    """

    
    def __init__(self,filename):
        self.systemBusData = {}
        self.filename = filename

        self.openWorksheet()


    def openWorksheet(self):
        if self.filename:
            self.wb = openpyxl.load_workbook(filename = self.filename)
            self.ws = self.wb.get_active_sheet()
            print 'rostertable: opened sheet: %s' % self.ws.title

        else:
            raise Exception('rostertable: no filename')


    def extractroster(self):
        cell2int = lambda s: int(filter( lambda x: x in '0123456789', str(s))) \
                if (s and filter( lambda x: x in '0123456789', str(s))) \
                else None
        isSystemID = lambda i: i in range(700,999)
        isBusID = lambda b: (b in range(1,699)) or (b in range(1000,10000))
        #test if cell value is a string  [convoluted make more simple]
        isString = lambda s: True if len(s if isinstance(s,basestring) else '') else False

        if self.ws:
            #maxrow = self.ws.get_hightest_row()  # no get_highest_row function!!!
            maxrow = cell2int( self.ws.calculate_dimension().split(':')[1] )

            _currentSystemID = None
            for row in range(0,maxrow):
                c = cell2int( self.ws.cell(column=0,row=row).value )
                textLabel = isString( self.ws.cell(column=1,row=row).value )

                if isSystemID( c ) and textLabel: _currentSystemID = c
                elif isBusID( c ):
                    if _currentSystemID: 
                        try:
                            self.systemBusData[_currentSystemID].append(c)
                        except KeyError:
                            self.systemBusData[_currentSystemID] = [c]
                    else:
                        print 'rostertable: extractroster: ERROR bus id with no system id (row %s)' % str(row)
        else:
            raise Exception('rostertable: ws variable not set')
    

class gfiData():
    gfiSystems = [
        [1,'Victoria'],
        [2,'Langford (Victoria)'],
        [3,'Whistler'],
        [4,'Squamish'],
        [5,'Nanaimo'],
        [6,'Abbotsford'],
        [7,'Kelowna'],
        [8,'Kamloops'],
        [9,'Prince George'],
        [10,'Cowichan Valley Commuter & Cowichan Conventional'],
        [11,'Trail'],
        [12,'Comox'],
        [13,'Port Alberni'],
        [14,'Campbell River'],
        [15,'Powell River'],
        [16,'Sunshine'],
        [17,'Vernon'],
        [18,'Penticton'],
        [19,'Chilliwack'],
        [20,'Cranbrook'],
        [21,'Nelson'],
        [22,'Terrace'],
        [23,'Prince Rupert'],
        [24,'Kitimat'],
        [25,'Fort St John']
        ]

    accountingSystems = [
        [701,'AGASSIZ-HARRISON PARA'],
        [804,'CAMPBELL RIVER CONVENTIONAL'],
        [801,'CENTRAL FRASER VALLEY CONV'],
        [811,'CHILLIWACK CONVENTIONAL'],
        [812,'COMOX CONVENTIONAL'],
        [813,'COWICHAN CONVENTIONAL'],
        [805,'COWICHAN VALLEY COMMUTER'],
        [810,'CRANBROOK CONVENTIONAL'],
        [814,'DAWSON CREEK CONVENTIONAL'],
        [817,'FORT ST. JOHN CONVENTIONAL'],
        [821,'KAMLOOPS CONVENTIONAL'],
        [824,'KELOWNA CONVENTIONAL'],
        [827,'KITIMAT CONVENTIONAL'],
        [841,'NANAIMO CONVENTIONAL'],
        [844,'NELSON CONVENTIONAL'],
        [846,'North Okanagan Connector'],
        [748,'PEMBERTON VALLEY (SLRD) PARA'],
        [847,'PENTICTON CONVENTIONAL'],
        [851,'PORT ALBERNI CONVENTIONAL'],
        [735,'PORT EDWARD PARA'],
        [854,'POWELL RIVER CONVENTIONAL'],
        [857,'PRINCE GEORGE CONVENTIONAL'],
        [861,'PRINCE RUPERT CONVENTIONAL'],
        [763,'QUESNEL PARA'],
        [771,'SKEENA REGIONAL PARA TRANSIT'],
        [865,'SQUAMISH CONVENTIONAL'],
        [866,'SUNSHINE COAST CONVENTIONAL'],
        [867,'TERRACE CONVENTIONAL'],
        [831,'TRAIL CONVENTIONAL'],
        [871,'VERNON CONVENTIONAL'],
        [990,'VICTORIA CONVENTIONAL'],
        [875,'WHISTLER CONVENTIONAL'],
        [772,'WILLIAMS LAKE PARA']
        ]

    systemMatch = [
        [990,1],
        [990,2],
        [875,3],
        [865,4],
        [841,5],
        [801,6],
        [824,7],
        [821,8],
        [857,9],
        [805,10],
        [813,10],
        [831,11],
        [812,12],
        [851,13],
        [804,14],
        [854,15],
        [866,16],
        [846,17],
        [871,17],
        [847,18],
        [811,19],
        [810,20],
        [844,21],
        [867,22],
        [861,23],
        [827,24],
        [817,25]
        ]



def open_sqlite( filename,rewrite ):
    addtables = True
    if os.path.isfile(filename):
        if rewrite: os.remove( filename )
        else: addtables = False
    sqliteconnection = sqlite3.connect( filename )
    if addtables: writebasetables_sqlite(sqliteconnection) 

    return sqliteconnection


    
def writebasetables_sqlite( sqliteconnection ):
    sqlitecursor = sqliteconnection.cursor()
    sqlitecursor.execute("CREATE TABLE gfisystems(gfiid INT,gfiname TEXT)")
    sqlitecursor.executemany("INSERT INTO gfisystems VALUES(?,?)",gfiData.gfiSystems)
    sqliteconnection.commit() 

    sqlitecursor.execute("CREATE TABLE accountingsystems(accountingcode INT,accountingname TEXT)")
    sqlitecursor.executemany("INSERT INTO accountingsystems VALUES(?,?)",gfiData.accountingSystems)
    sqliteconnection.commit() 

    sqlitecursor.execute("CREATE TABLE systemmatch(accountingcode INT,gfiid INT)")
    sqlitecursor.executemany("INSERT INTO systemmatch VALUES(?,?)",gfiData.systemMatch)
    sqliteconnection.commit() 

    sqlitecursor.execute("CREATE TABLE busroster(accountingcode INT,busid INT,month TEXT)")
    sqliteconnection.commit() 


    
def write_sqlite( sqliteconnection, busdata, month):
    dbdata = []
    sqlitecursor = sqliteconnection.cursor()
    for s in busdata:
        for b in busdata[s]:
            dbdata.append([s,b,month])
    sqlitecursor.executemany("INSERT INTO busroster VALUES(?,?,?)",dbdata)
    sqliteconnection.commit() 


    
def close_sqlite( sqliteconnection ):
    sqliteconnection.close()



class rosterdb():
    """
    Class representing database and table(s) version of bus roster
     - the output.
    """


    filetypeActions = {
        'db': {'open':open_sqlite,'write':write_sqlite,'close':close_sqlite}
    }


    def __init__(self,filename=None,rewrite=None,connectionstring=None):
        if filename: 
            self.filename = filename
            if len(filename.split('.')) > 1:
                self.fileType =  filename.split('.')[-1].strip().lower()
            else:
                raise Exception('rosterdb: missing file extension')
        if connectionstring: self.connectionstring = connectionstring

        self.rewrite = False
        if rewrite: self.rewrite = rewrite

        if self.filename: self.opendb()


    def opendb(self,filename=None,rewrite=None,connectionstring=None):
        if filename: 
            self.filename = filename
            if len(filename.split('.')) > 1:
                self.fileType =  filename.split('.')[-1].strip().lower()
            else:
                raise Exception('rosterdb: missing file extension')
        if connectionstring: self.connectionstring = connectionstring
        if rewrite: self.rewrite = rewrite

        if self.filename and self.fileType:
            if self.fileType in self.filetypeActions.keys():
                self.connection = self.filetypeActions[self.fileType]['open'](self.filename, self.rewrite)
            else:
                raise Exception('rosterdb: unknown file extension')
        else:
            raise Exception('rosterdb: filename not set')

    
    def writedb(self,busdata,month):
        if self.filename and self.fileType in self.filetypeActions.keys():
                self.filetypeActions[self.fileType]['write'](self.connection,busdata,month)
        else:
            raise Exception('rosterdb: writedb: file not setup')


    def closedb(self):
        if self.filename and self.fileType in self.filetypeActions.keys():
                self.filetypeActions[self.fileType]['close'](self.connection)
        else:
            raise Exception('rosterdb: closedb: file not setup')
        



def getArgs():
    argsPsr = argparse.ArgumentParser(description='Insert given bus roster into db')
    argsPsr.add_argument('-d','--database',required=True,help='output database')
    argsPsr.add_argument('-f','--file',required=True,help='input xlsx file')
    argsPsr.add_argument('-m','--month',required=True,help='eg: 2014-05')
    argsPsr.add_argument('-r','--rewrite',required=False,action='store_true',help='remove tables/db before adding')
    argsPsr.add_argument('-c','--connection',required=False,help='connection string')
    args = argsPsr.parse_args()
    return args


if __name__ == '__main__':

    args = getArgs()
    
    rt = rostertable(args.file)
    rt.extractroster()

    r = rosterdb(filename=args.database,rewrite=args.rewrite)
    r.writedb(rt.systemBusData,args.month)
    r.closedb()

    print "Completed."
    sys.exit(0)


